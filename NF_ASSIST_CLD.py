# ================================================================
# ASSISTENTE FISCAL NF-e / NFS-e - TECNOGLOBO EQUIPAMENTOS LTDA
# ================================================================
# Instale as bibliotecas necessarias:
# pip install openpyxl fpdf2 requests
# (weasyprint foi substituido por fpdf2 para evitar problemas de dependencia no Windows)
# ================================================================

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import os, sys
import webbrowser
import json

# --- Verificacao de bibliotecas ---
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from fpdf import FPDF # Usando fpdf2
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

# ================================================================
# DADOS DO EMITENTE - ALTERE APENAS AQUI
# ================================================================
EMITENTE = {
    "razao_social":       "TECNOGLOBO EQUIPAMENTOS LTDA",
    "nome_fantasia":      "TECNOGLOBO",
    "cnpj":               "07.700.041/0001-84",
    "inscricao_estadual": "9035777625",
    "inscricao_municipal":"",
    "endereco":           "Rua Joao David Perneta, 355",
    "bairro":             "Hugo Langue",
    "cidade":             "Curitiba",
    "uf":                 "PR",
    "cep":                "80040-330",
    "telefone":           "",
    "email":              "",
    "regime_tributario":  "Lucro Presumido",
    "crt":                "3",
}

# ================================================================
# PALETA DE CORES TECNOGLOBO (Laranja e Marrom)
# ================================================================
CORES = {
    "PRIMARIA": "#E65100",  # Laranja Escuro
    "SECUNDARIA": "#5D4037", # Marrom Escuro
    "ACENTO": "#FFB74D",    # Laranja Claro
    "TEXTO_CLARO": "#FFFFFF",
    "TEXTO_ESCURO": "#3E2723", # Marrom Muito Escuro
    "FUNDO_CLARO": "#F5F5F5",
    "FUNDO_MEDIO": "#EFEBE9", # Marrom Claro
    "FUNDO_ALERTA": "#FFF8E1", # Amarelo Claro
    "BORDA": "#D7CCC8",     # Marrom Acinzentado
    "SUCESSO": "#388E3C",   # Verde para sucesso
    "ERRO": "#D32F2F",      # Vermelho para erro
    "AVISO": "#FFA000",     # Laranja para aviso
}

# ================================================================
# MAPEAMENTO NCM -> CFOP (Exemplo - ADICIONE MAIS CONFORME NECESSARIO)
# ================================================================
NCM_CFOP_MAP = {
    "84713019": ["5.102 - Venda dentro do estado", "6.102 - Venda para outro estado"],
    "84433199": ["5.102 - Venda dentro do estado", "6.102 - Venda para outro estado"],
    "85176299": ["5.102 - Venda dentro do estado", "6.102 - Venda para outro estado"],
    "85287200": ["5.102 - Venda dentro do estado", "6.102 - Venda para outro estado"],
    "84714900": ["5.102 - Venda dentro do estado", "6.102 - Venda para outro estado"],
    "5.949":    ["5.949 - Remessa demonstracao no estado"],
    "6.949":    ["6.949 - Remessa demonstracao outro estado"],
    "5.411":    ["5.411 - Devolucao no estado"],
    "6.411":    ["6.411 - Devolucao para outro estado"],
}

# ================================================================
# CLASSE GERADOR DANFE (AGORA NO ESCOPO GLOBAL)
# ================================================================
class GeradorDANFE:

    def __init__(self, dados):
        self.dados = dados
        self.pdf = FPDF(unit="mm", format="A4")
        self.pdf.set_auto_page_break(auto=True, margin=15)
        self.M = 14 # Margem
        self.L = 210 - (2 * self.M) # Largura util

    def _set_font(self, size, bold=False, italic=False, color=None):
        style = ""
        if bold: style += "B"
        if italic: style += "I"
        self.pdf.set_font("Helvetica", style, size)
        if color:
            if isinstance(color, str) and color.startswith("#"):
                color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
            self.pdf.set_text_color(*color)
        else:
            self.pdf.set_text_color(0, 0, 0) # Preto padrao

    def _rect_fill(self, x, y, w, h, color):
        if isinstance(color, str) and color.startswith("#"):
            color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
        self.pdf.set_fill_color(*color)
        self.pdf.rect(x, y, w, h, style="F")

    def _rect_border(self, x, y, w, h, color=CORES["BORDA"], thickness=0.2):
        if isinstance(color, str) and color.startswith("#"):
            color = tuple(int(color[i:i+2], 16) for i in (1, 3, 5))
        self.pdf.set_draw_color(*color)
        self.pdf.set_line_width(thickness)
        self.pdf.rect(x, y, w, h, style="D")

    def _titulo_sec(self, titulo, cor=CORES["PRIMARIA"]):
        p = self.pdf
        p.ln(2)
        y = p.get_y()
        self._rect_fill(self.M, y, self.L, 7, cor)
        self._rect_border(self.M, y, self.L, 7)
        self._set_font(9, bold=True, color=CORES["TEXTO_CLARO"])
        p.set_xy(self.M + 2, y + 1.5)
        p.cell(self.L - 4, 4, titulo, 0, 1, "L")
        p.ln(1)

    def _linha_dado(self, campo, valor, alt=False):
        p = self.pdf
        y = p.get_y()
        bg_color = CORES["FUNDO_MEDIO"] if alt else CORES["FUNDO_CLARO"]
        self._rect_fill(self.M, y, self.L, 6, bg_color)
        self._rect_border(self.M, y, self.L, 6)

        self._set_font(8, bold=True, color=CORES["SECUNDARIA"])
        p.set_xy(self.M + 2, y + 1)
        p.cell(self.L * 0.3, 4, campo, 0, 0, "L")

        self._set_font(8, color=CORES["TEXTO_ESCURO"])
        p.cell(self.L * 0.7 - 4, 4, valor, 0, 1, "L")
        p.ln(0)

    def gerar(self, caminho):
        self.pdf.add_page()
        d = self.dados
        dh = datetime.now().strftime("%d/%m/%Y %H:%M")
        dh_full = datetime.now().strftime("%d/%m/%Y as %H:%M:%S")

        # --- CABECALHO ---
        y_cab = self.pdf.get_y()
        self._rect_fill(self.M, y_cab, self.L, 20, CORES["SECUNDARIA"])
        self._rect_border(self.M, y_cab, self.L, 20, CORES["SECUNDARIA"], 2)

        # Emitente
        self._set_font(12, bold=True, color=CORES["TEXTO_CLARO"])
        self.pdf.set_xy(self.M + 2, y_cab + 2)
        self.pdf.cell(self.L * 0.52, 5, EMITENTE['razao_social'], 0, 1, "L")
        self._set_font(7, color=CORES["ACENTO"])
        self.pdf.set_x(self.M + 2)
        self.pdf.multi_cell(self.L * 0.52, 3,
                            (f"{EMITENTE['endereco']} - {EMITENTE['bairro']} - "
                             f"{EMITENTE['cidade']}/{EMITENTE['uf']} - CEP {EMITENTE['cep']}\n"
                             f"CNPJ: {EMITENTE['cnpj']}  |  IE: {EMITENTE['inscricao_estadual']}  |  "
                             f"Regime: {EMITENTE['regime_tributario']}"), 0, "L")

        # DANFE
        self._rect_fill(self.M + self.L * 0.52, y_cab, self.L * 0.24, 20, CORES["SECUNDARIA"])
        self._rect_border(self.M + self.L * 0.52, y_cab, self.L * 0.24, 20, CORES["ACENTO"], 0.2)
        self._set_font(13, bold=True, color=CORES["TEXTO_CLARO"])
        self.pdf.set_xy(self.M + self.L * 0.52, y_cab + 2)
        self.pdf.cell(self.L * 0.24, 5, "DANFE", 0, 1, "C")
        self._set_font(7, color=CORES["ACENTO"])
        self.pdf.set_x(self.M + self.L * 0.52)
        self.pdf.multi_cell(self.L * 0.24, 3, "Documento Auxiliar da\nNota Fiscal Eletronica", 0, "C")
        self._set_font(8, bold=True, color=CORES["TEXTO_CLARO"])
        self.pdf.set_xy(self.M + self.L * 0.52, y_cab + 14)
        self.pdf.cell(self.L * 0.24, 4, "ENTRADA 1 / SAIDA 2", 0, 1, "C")

        # Numeracao
        self._rect_fill(self.M + self.L * 0.76, y_cab, self.L * 0.24, 20, CORES["FUNDO_MEDIO"])
        self._rect_border(self.M + self.L * 0.76, y_cab, self.L * 0.24, 20, CORES["BORDA"], 0.2)
        self.pdf.set_xy(self.M + self.L * 0.76 + 2, y_cab + 2)
        self._set_font(7, bold=True, color=CORES["SECUNDARIA"])
        self.pdf.cell(self.L * 0.1, 4, "No. NF:", 0, 0, "L")
        self._set_font(9, bold=True, color=CORES["ERRO"])
        self.pdf.cell(self.L * 0.14 - 4, 4, "000.000.000", 0, 1, "L")

        self.pdf.set_x(self.M + self.L * 0.76 + 2)
        self._set_font(7, bold=True, color=CORES["SECUNDARIA"])
        self.pdf.cell(self.L * 0.1, 4, "Serie:", 0, 0, "L")
        self._set_font(8, color=CORES["TEXTO_ESCURO"])
        self.pdf.cell(self.L * 0.14 - 4, 4, "001", 0, 1, "L")

        self.pdf.set_x(self.M + self.L * 0.76 + 2)
        self._set_font(7, bold=True, color=CORES["SECUNDARIA"])
        self.pdf.cell(self.L * 0.1, 4, "Emissao:", 0, 0, "L")
        self._set_font(8, color=CORES["TEXTO_ESCURO"])
        self.pdf.cell(self.L * 0.14 - 4, 4, dh, 0, 1, "L")

        self.pdf.set_x(self.M + self.L * 0.76 + 2)
        self._set_font(7, bold=True, color=CORES["SECUNDARIA"])
        self.pdf.cell(self.L * 0.1, 4, "CRT:", 0, 0, "L")
        self._set_font(8, color=CORES["TEXTO_ESCURO"])
        self.pdf.cell(self.L * 0.14 - 4, 4, EMITENTE['crt'], 0, 1, "L")
        self.pdf.ln(2)

        # --- CHAVE DE ACESSO ---
        y_chave = self.pdf.get_y()
        self._rect_fill(self.M, y_chave, self.L, 8, CORES["FUNDO_MEDIO"])
        self._rect_border(self.M, y_chave, self.L, 8, CORES["SECUNDARIA"], 1.5)
        self._set_font(8, bold=True, color=CORES["SECUNDARIA"])
        self.pdf.set_xy(self.M + 2, y_chave + 2)
        self.pdf.cell(self.L * 0.3, 4, "CHAVE DE ACESSO:", 0, 0, "L")
        self._set_font(8, color=CORES["SECUNDARIA"])
        self.pdf.cell(self.L * 0.7 - 4, 4, "0000 0000 0000 0000   0000 0000 0000 0000   0000 0000 0000", 0, 1, "C")
        self.pdf.ln(2)

        # --- AVISO ---
        y_aviso = self.pdf.get_y()
        self._rect_fill(self.M, y_aviso, self.L, 12, CORES["FUNDO_ALERTA"])
        self._rect_border(self.M, y_aviso, self.L, 12, CORES["AVISO"], 2)
        self._set_font(8, bold=True, color=CORES["AVISO"])
        self.pdf.set_xy(self.M + 2, y_aviso + 2)
        self.pdf.multi_cell(self.L - 4, 4,
                            "⚠ DOCUMENTO SEM VALOR FISCAL ⚠\n"
                            "Este espelho e apenas um RASCUNHO para conferencia interna. "
                            "Nao substitui a NF-e oficial transmitida e autorizada pela SEFAZ. "
                            "Confirme NCM e CFOP com seu contador.", 0, "C")
        self.pdf.ln(2)

        # --- NATUREZA ---
        self._titulo_sec("NATUREZA DA OPERACAO")
        self._linha_dado("Tipo de Nota", d.get('tipo_nota','-'), False)
        self._linha_dado("Natureza", d.get('natureza','-'), True)
        self._linha_dado("CFOP", d.get('cfop','-'), False)
        self._linha_dado("NCM", d.get('ncm','-'), True)

        # --- DESTINATARIO ---
        self._titulo_sec("DESTINATARIO")
        self._linha_dado("Razao Social / Nome", d.get('razao_social','-'), False)
        self._linha_dado("CPF / CNPJ", d.get('cpf_cnpj','-'), True)
        self._linha_dado("Inscricao Estadual", d.get('insc_estadual','-'), False)
        self._linha_dado("Endereco", d.get('endereco','-'), True)
        self._linha_dado("E-mail", d.get('email','-'), False)

        # --- PRODUTO ---
        self._titulo_sec("DADOS DO PRODUTO / EQUIPAMENTO")
        p = self.pdf
        p.ln(1)
        y_prod_header = p.get_y()
        self._rect_fill(self.M, y_prod_header, self.L, 7, CORES["SECUNDARIA"])
        self._rect_border(self.M, y_prod_header, self.L, 7, CORES["ACENTO"])
        self._set_font(7, bold=True, color=CORES["TEXTO_CLARO"])
        p.set_xy(self.M, y_prod_header + 1.5)
        p.cell(self.L * 0.05, 4, "Item", 0, 0, "C")
        p.cell(self.L * 0.33, 4, "Descricao", 0, 0, "C")
        p.cell(self.L * 0.12, 4, "NCM", 0, 0, "C")
        p.cell(self.L * 0.07, 4, "Qtd", 0, 0, "C")
        p.cell(self.L * 0.13, 4, "Vlr Unit", 0, 0, "C")
        p.cell(self.L * 0.15, 4, "Vlr Total", 0, 0, "C")
        p.cell(self.L * 0.15, 4, "N. Serie", 0, 1, "C")

        y_prod_data = p.get_y()
        self._rect_fill(self.M, y_prod_data, self.L, 10, CORES["FUNDO_MEDIO"])
        self._rect_border(self.M, y_prod_data, self.L, 10, CORES["BORDA"])
        self._set_font(8, color=CORES["TEXTO_ESCURO"])
        p.set_xy(self.M, y_prod_data + 1)
        p.cell(self.L * 0.05, 8, "001", 0, 0, "C")
        p.cell(self.L * 0.33, 8, d.get('descricao','-'), 0, 0, "L")
        p.cell(self.L * 0.12, 8, d.get('ncm','-'), 0, 0, "C")
        p.cell(self.L * 0.07, 8, d.get('quantidade','-'), 0, 0, "C")
        p.cell(self.L * 0.13, 8, d.get('valor_unit','-'), 0, 0, "C")
        self._set_font(9, bold=True, color=CORES["SUCESSO"])
        p.cell(self.L * 0.15, 8, d.get('valor_total','-'), 0, 0, "C")
        self._set_font(8, color=CORES["TEXTO_ESCURO"])
        p.cell(self.L * 0.15, 8, d.get('num_serie','-'), 0, 1, "C")
        p.ln(2)

        # --- TOTAIS ---
        self._titulo_sec("CALCULO DO IMPOSTO E TOTAIS")
        p = self.pdf
        ni = "A calcular no sistema"
        raw = [
            ("BC ICMS",         ni,    "Valor ICMS",    ni),
            ("BC ICMS ST",      ni,    "Valor ICMS ST", ni),
            ("Valor Produtos",  d.get('valor_total','-'),    "Valor Frete",   "Verificar"),
            ("Valor Seguro",   "0,00", "Desconto",      "0,00"),
            ("Outras Despesas","0,00", "Valor IPI",     ni),
            ("Valor PIS",       ni,    "Valor COFINS",  ni),
            ("VALOR TOTAL NF",  d.get('valor_total','-'),    "Forma Pgto",    d.get("pagamento","-")),
        ]
        for i, (l1, v1, l2, v2) in enumerate(raw):
            y = p.get_y()
            is_total_row = (l1 == "VALOR TOTAL NF")
            bg_color = CORES["SUCESSO"] if is_total_row else (CORES["FUNDO_MEDIO"] if i % 2 == 0 else CORES["FUNDO_CLARO"])
            text_color = CORES["TEXTO_CLARO"] if is_total_row else CORES["TEXTO_ESCURO"]

            self._rect_fill(self.M, y, self.L, 6, bg_color)
            self._rect_border(self.M, y, self.L, 6)

            self._set_font(8, bold=True, color=CORES["SECUNDARIA"])
            p.set_xy(self.M + 2, y + 1)
            p.cell(self.L * 0.25, 4, l1, 0, 0, "L")

            self._set_font(8, bold=is_total_row, color=text_color)
            p.cell(self.L * 0.25 - 4, 4, v1, 0, 0, "L")

            self._set_font(8, bold=True, color=CORES["SECUNDARIA"])
            p.cell(self.L * 0.25, 4, l2, 0, 0, "L")

            self._set_font(8, bold=is_total_row, color=text_color)
            p.cell(self.L * 0.25 - 4, 4, v2, 0, 1, "L")
            p.ln(0)

        p.ln(1)
        self._set_font(7, italic=True, color=CORES["TEXTO_ESCURO"])
        p.set_x(self.M)
        p.multi_cell(self.L, 3,
                     "* ICMS, IPI, PIS, COFINS e ICMS-ST devem ser calculados pelo sistema "
                     "de emissao conforme o regime tributario e a UF do destinatario.", 0, "L")
        p.ln(2)

        self._titulo_sec("TRANSPORTE")
        self._linha_dado("Modalidade do Frete", d.get('frete','-'), False)
        self._linha_dado("Transportadora", d.get('transportadora','-'), True)
        self._linha_dado("Qtd / Especie", "1 / EQUIPAMENTO", False)

        self._titulo_sec("DADOS ADICIONAIS / OBSERVACOES")
        self._linha_dado("Observacoes", d.get('observacoes','-'), False)

        # --- PROXIMOS PASSOS ---
        self._titulo_sec("PROXIMOS PASSOS PARA EMISSAO OFICIAL", CORES["SUCESSO"])
        passos = [
            "1. Confirme o NCM e o CFOP com seu contador ou departamento fiscal.",
            "2. Acesse seu sistema de emissao (ERP, emissor SEFAZ ou plataforma NF-e).",
            "3. Preencha os dados conforme este rascunho.",
            "4. Calcule os impostos (ICMS, PIS, COFINS, IPI) com auxilio do sistema.",
            "5. Revise todos os dados e transmita a nota para a SEFAZ.",
            "6. Aguarde a autorizacao e envie o XML e o DANFE oficial ao destinatario.",
        ]
        for i, passo in enumerate(passos):
            y = p.get_y()
            bg_color = CORES["FUNDO_MEDIO"] if i % 2 == 0 else CORES["FUNDO_CLARO"]
            self._rect_fill(self.M, y, self.L, 6, bg_color)
            self._rect_border(self.M, y, self.L, 6)
            self._set_font(8, color=CORES["SECUNDARIA"])
            p.set_xy(self.M + 2, y + 1)
            p.multi_cell(self.L - 4, 4, passo, 0, "L")
            p.ln(0)
        p.ln(2)

        # --- RODAPE ---
        y_rodape = p.get_y()
        self._rect_fill(self.M, y_rodape, self.L, 6, CORES["FUNDO_MEDIO"])
        self._rect_border(self.M, y_rodape, self.L, 6)
        self._set_font(7, italic=True, color=CORES["TEXTO_ESCURO"])
        self.pdf.set_xy(self.M, y_rodape + 1)
        self.pdf.cell(self.L, 4,
               f"Gerado em {dh_full}  |  {EMITENTE['razao_social']}  |  "
               f"CNPJ: {EMITENTE['cnpj']}  |  DOCUMENTO SEM VALOR FISCAL", 0, 1, "C")

        self.pdf.output(caminho)

    # --- Metodo para gerar HTML (mantido para exportacao HTML) ---
    def _montar_html(self):
        d = self.dados
        dh = datetime.now().strftime("%d/%m/%Y %H:%M")
        dh_full = datetime.now().strftime("%d/%m/%Y as %H:%M:%S")

        try:
            qtd = float(d.get("quantidade","1").replace(",","."))
            vlr = float(d.get("valor_unit","0")
                        .replace("R$","").replace(".","").replace(",",".").strip())
            tot = qtd * vlr
            ts  = f"R$ {tot:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            vs  = f"R$ {vlr:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            ts = "Verificar"; vs = d.get("valor_unit","-")

        ni = "A calcular no sistema"

        totais_rows = ""
        raw = [
            ("BC ICMS",         ni,    "Valor ICMS",    ni),
            ("BC ICMS ST",      ni,    "Valor ICMS ST", ni),
            ("Valor Produtos",  ts,    "Valor Frete",   "Verificar"),
            ("Valor Seguro",   "0,00", "Desconto",      "0,00"),
            ("Outras Despesas","0,00", "Valor IPI",     ni),
            ("Valor PIS",       ni,    "Valor COFINS",  ni),
            ("VALOR TOTAL NF",  ts,    "Forma Pgto",    d.get("pagamento","-")),
        ]
        for i,(l1,v1,l2,v2) in enumerate(raw):
            is_tot = (i == len(raw)-1)
            bg     = CORES["SUCESSO"] if is_tot else (CORES["FUNDO_MEDIO"] if i%2==0 else CORES["FUNDO_CLARO"])
            fw     = "bold"    if is_tot else "normal"
            cor_v  = CORES["TEXTO_CLARO"] if is_tot else CORES["TEXTO_ESCURO"]
            totais_rows += f"""
            <tr style="background:{bg}">
                <td class="campo">{l1}</td>
                <td class="valor" style="font-weight:{fw};color:{cor_v}">{v1}</td>
                <td class="campo">{l2}</td>
                <td class="valor" style="font-weight:{fw};color:{cor_v}">{v2}</td>
            </tr>"""

        passos_html = ""
        passos = [
            "1. Confirme o NCM e o CFOP com seu contador ou departamento fiscal.",
            "2. Acesse seu sistema de emissao (ERP, emissor SEFAZ ou plataforma NF-e).",
            "3. Preencha os dados conforme este rascunho.",
            "4. Calcule os impostos (ICMS, PIS, COFINS, IPI) com auxilio do sistema.",
            "5. Revise todos os dados e transmita a nota para a SEFAZ.",
            "6. Aguarde a autorizacao e envie o XML e o DANFE oficial ao destinatario.",
        ]
        for i, p in enumerate(passos):
            bg = CORES["FUNDO_MEDIO"] if i%2==0 else CORES["FUNDO_CLARO"]
            passos_html += f'<tr style="background:{bg}"><td class="passo">{p}</td></tr>'

        return f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<style>
  @page {{
    size: A4;
    margin: 14mm;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; font-family: Arial, sans-serif; }}

  body::before {{
    content: "SEM VALOR FISCAL";
    position: fixed;
    top: 38%;
    left: 5%;
    font-size: 58px;
    font-weight: bold;
    color: rgba(180,180,180,0.18);
    transform: rotate(45deg);
    white-space: nowrap;
    z-index: 0;
  }}

  body {{ font-size: 9px; color: {CORES["TEXTO_ESCURO"]}; }}

  .cab-wrap {{
    display: flex;
    border: 2px solid {CORES["SECUNDARIA"]};
    margin-bottom: 3px;
  }}
  .cab-emit {{
    background: {CORES["SECUNDARIA"]};
    color: {CORES["TEXTO_CLARO"]};
    padding: 6px 8px;
    flex: 0 0 52%;
    display: flex;
    flex-direction: column;
    justify-content: center;
  }}
  .cab-emit .razao  {{ font-size:12px; font-weight:bold; margin-bottom:3px; }}
  .cab-emit .info   {{ font-size:7px; color:{CORES["ACENTO"]}; line-height:1.5; }}
  .cab-danfe {{
    background: {CORES["SECUNDARIA"]};
    color: {CORES["TEXTO_CLARO"]};
    flex: 0 0 24%;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    border-left: 1px solid {CORES["ACENTO"]};
    border-right: 1px solid {CORES["ACENTO"]};
    padding: 4px;
    text-align: center;
  }}
  .cab-danfe .d-title {{ font-size:13px; font-weight:bold; }}
  .cab-danfe .d-sub   {{ font-size:7px; color:{CORES["ACENTO"]}; margin-top:2px; }}
  .cab-danfe .d-es    {{ font-size:8px; font-weight:bold; margin-top:4px; }}
  .cab-num {{
    background: {CORES["FUNDO_MEDIO"]};
    flex: 0 0 24%;
    padding: 4px 6px;
    font-size: 7px;
  }}
  .cab-num table {{ width:100%; border-collapse:collapse; }}
  .cab-num td {{ padding: 2px 3px; border-bottom: 1px solid {CORES["BORDA"]}; }}
  .cab-num .nl {{ font-weight:bold; color:{CORES["SECUNDARIA"]}; }}
  .cab-num .nv {{ color:{CORES["ERRO"]}; font-weight:bold; font-size:9px; }}

  .chave {{
    background: {CORES["FUNDO_MEDIO"]};
    border: 1.5px solid {CORES["SECUNDARIA"]};
    padding: 4px 8px;
    font-size: 8px;
    color: {CORES["SECUNDARIA"]};
    font-family: Courier, monospace;
    text-align: center;
    margin-bottom: 4px;
  }}
  .chave span {{ font-weight:bold; color:{CORES["SECUNDARIA"]}; margin-right:8px; }}

  .aviso {{
    background: {CORES["FUNDO_ALERTA"]};
    border: 2px solid {CORES["AVISO"]};
    padding: 5px 10px;
    font-size: 8px;
    font-weight: bold;
    color: {CORES["AVISO"]};
    text-align: center;
    margin-bottom: 5px;
  }}

  table.dados {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 5px;
    font-size: 8px;
  }}
  .titulo-sec {{
    background: {CORES["PRIMARIA"]};
    color: {CORES["TEXTO_CLARO"]};
    font-weight: bold;
    font-size: 9px;
    padding: 4px 8px;
  }}
  td.campo {{
    font-weight: bold;
    color: {CORES["SECUNDARIA"]};
    width: 28%;
    padding: 3px 6px;
    border: 1px solid {CORES["BORDA"]};
    vertical-align: top;
  }}
  td.valor {{
    color: {CORES["TEXTO_ESCURO"]};
    padding: 3px 6px;
    border: 1px solid {CORES["BORDA"]};
    vertical-align: top;
  }}
  td.valor.total {{
    color: {CORES["SUCESSO"]};
    font-weight: bold;
    font-size: 10px;
  }}

  table.produtos {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 5px;
    font-size: 8px;
  }}
  table.produtos th {{
    background: {CORES["SECUNDARIA"]};
    color: {CORES["TEXTO_CLARO"]};
    padding: 3px 4px;
    text-align: center;
    font-size: 7px;
    border: 1px solid {CORES["ACENTO"]};
  }}
  table.produtos td {{
    background: {CORES["FUNDO_MEDIO"]};
    padding: 4px;
    text-align: center;
    border: 1px solid {CORES["BORDA"]};
  }}
  .vlr-total-prod {{ color:{CORES["SUCESSO"]}; font-weight:bold; font-size:9px; }}

  table.totais {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 3px;
    font-size: 8px;
  }}
  table.totais td {{ padding: 3px 6px; border: 1px solid {CORES["BORDA"]}; }}
  .nota-imp {{
    font-size: 7px;
    color: {CORES["TEXTO_ESCURO"]};
    font-style: italic;
    margin-bottom: 5px;
    padding: 3px 6px;
    background: {CORES["FUNDO_CLARO"]};
    border: 1px solid {CORES["BORDA"]};
  }}

  table.passos {{
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 5px;
    font-size: 8px;
  }}
  td.passo {{
    padding: 4px 10px;
    border: 1px solid {CORES["BORDA"]};
    color: {CORES["SECUNDARIA"]};
  }}

  .rodape {{
    background: {CORES["FUNDO_MEDIO"]};
    border: 1px solid {CORES["BORDA"]};
    padding: 4px 8px;
    font-size: 7px;
    color: {CORES["TEXTO_ESCURO"]};
    text-align: center;
    margin-top: 6px;
  }}
</style>
</head>
<body>

<!-- CABECALHO -->
<div class="cab-wrap">
  <div class="cab-emit">
    <div class="razao">{EMITENTE['razao_social']}</div>
    <div class="info">
      {EMITENTE['endereco']} - {EMITENTE['bairro']} -
      {EMITENTE['cidade']}/{EMITENTE['uf']} - CEP {EMITENTE['cep']}<br>
      CNPJ: {EMITENTE['cnpj']}  |  IE: {EMITENTE['inscricao_estadual']}  |
      Regime: {EMITENTE['regime_tributario']}
    </div>
  </div>
  <div class="cab-danfe">
    <div class="d-title">DANFE</div>
    <div class="d-sub">Documento Auxiliar da<br>Nota Fiscal Eletronica</div>
  </div>
  <div class="cab-num">
    <table>
      <tr><td class="nl">No. NF:</td>  <td class="nv">000.000.000</td></tr>
      <tr><td class="nl">Serie:</td>   <td>001</td></tr>
      <tr><td class="nl">Emissao:</td> <td>{dh}</td></tr>
      <tr><td class="nl">CRT:</td>     <td>{EMITENTE['crt']}</td></tr>
    </table>
  </div>
</div>

<!-- CHAVE -->
<div class="chave">
  <span>CHAVE DE ACESSO:</span>
  0000 0000 0000 0000 &nbsp; 0000 0000 0000 0000 &nbsp; 0000 0000 0000
</div>

<!-- AVISO -->
<div class="aviso">
  &#9888; DOCUMENTO SEM VALOR FISCAL &#9888;<br>
  Este espelho e apenas um RASCUNHO para conferencia interna.
  Nao substitui a NF-e oficial transmitida e autorizada pela SEFAZ.
  Confirme NCM e CFOP com seu contador.
</div>

<!-- NATUREZA -->
<table class="dados">
  {self._titulo_sec("NATUREZA DA OPERACAO")}
  {self._linha("Tipo de Nota",  d.get('tipo_nota','-'),  False)}
  {self._linha("Natureza",      d.get('natureza','-'),   True)}
  {self._linha("CFOP",          d.get('cfop','-'),       False)}
  {self._linha("NCM",           d.get('ncm','-'),        True)}
</table>

<!-- DESTINATARIO -->
<table class="dados">
  {self._titulo_sec("DESTINATARIO")}
  {self._linha("Razao Social / Nome",  d.get('razao_social','-'),  False)}
  {self._linha("CPF / CNPJ",          d.get('cpf_cnpj','-'),      True)}
  {self._linha("Inscricao Estadual",  d.get('insc_estadual','-'), False)}
  {self._linha("Endereco",            d.get('endereco','-'),      True)}
  {self._linha("E-mail",             d.get('email','-'),         False)}
</table>

<!-- PRODUTO -->
<table class="dados">
  {self._titulo_sec("DADOS DO PRODUTO / EQUIPAMENTO")}
</table>
<table class="produtos">
  <tr>
    <th style="width:5%">Item</th>
    <th style="width:33%">Descricao</th>
    <th style="width:12%">NCM</th>
    <th style="width:7%">Qtd</th>
    <th style="width:13%">Vlr Unit</th>
    <th style="width:15%">Vlr Total</th>
    <th style="width:15%">N. Serie</th>
  </tr>
  <tr>
    <td>001</td>
    <td style="text-align:left">{d.get('descricao','-')}</td>
    <td>{d.get('ncm','-')}</td>
    <td>{d.get('quantidade','-')}</td>
    <td>{vs}</td>
    <td class="vlr-total-prod">{ts}</td>
    <td>{d.get('num_serie','-')}</td>
  </tr>
</table>

<!-- TOTAIS -->
<table class="dados">
  {self._titulo_sec("CALCULO DO IMPOSTO E TOTAIS")}
</table>
<table class="totais">
  {totais_rows}
</table>
<div class="nota-imp">
  * ICMS, IPI, PIS, COFINS e ICMS-ST devem ser calculados pelo sistema
  de emissao conforme o regime tributario e a UF do destinatario.
</div>

<!-- TRANSPORTE -->
<table class="dados">
  {self._titulo_sec("TRANSPORTE")}
  {self._linha("Modalidade do Frete", d.get('frete','-'),          False)}
  {self._linha("Transportadora",      d.get('transportadora','-'), True)}
  {self._linha("Qtd / Especie",       "1 / EQUIPAMENTO",           False)}
</table>

<!-- OBSERVACOES -->
<table class="dados">
  {self._titulo_sec("DADOS ADICIONAIS / OBSERVACOES")}
</table>
<table class="passos">
  {self._linha("Observacoes", d.get('observacoes','-'), False)}
</table>

<!-- PROXIMOS PASSOS -->
<table class="dados">
  {self._titulo_sec("PROXIMOS PASSOS PARA EMISSAO OFICIAL",CORES["SUCESSO"])}
</table>
<table class="passos">
  {passos_html}
</table>

<!-- RODAPE -->
<div class="rodape">
  Gerado em {dh_full} &nbsp;|&nbsp; {EMITENTE['razao_social']} &nbsp;|&nbsp;
  CNPJ: {EMITENTE['cnpj']} &nbsp;|&nbsp; DOCUMENTO SEM VALOR FISCAL
</div>

</body>
</html>"""

# ================================================================
# ASSISTENTE FISCAL
# ================================================================
class AssistenteFiscal:

    DADOS_VAZIOS = {
        "tipo_nota":"", "razao_social":"", "cpf_cnpj":"",
        "insc_estadual":"", "endereco":"", "email":"",
        "descricao":"", "quantidade":"", "num_serie":"",
        "valor_unit":"", "valor_total":"", "ncm":"",
        "cfop":"", "natureza":"", "pagamento":"",
        "frete":"", "transportadora":"", "observacoes":"",
    }

    def __init__(self, root):
        self.root = root
        self.root.title(f"Assistente Fiscal - {EMITENTE['razao_social']}")
        self.root.geometry("820x700")
        self.root.resizable(True, True)
        self.root.configure(bg=CORES["FUNDO_CLARO"])
        self.dados = dict(self.DADOS_VAZIOS)
        self.etapa_atual = 0
        self.etapas = [
            self.etapa_boas_vindas,
            self.etapa_tipo_nota,
            self.etapa_destinatario,
            self.etapa_equipamento,
            self.etapa_codigos_fiscais,
            self.etapa_condicoes,
            self.etapa_observacoes,
            self.etapa_resumo_final,
        ]
        self._build_ui()
        self.etapa_boas_vindas()

    # --- UI Base ---
    def _build_ui(self):
        cab = tk.Frame(self.root, bg=CORES["SECUNDARIA"], pady=10)
        cab.pack(fill="x")
        tk.Label(cab, text=EMITENTE["razao_social"],
                 font=("Calibri",13,"bold"), fg=CORES["TEXTO_CLARO"], bg=CORES["SECUNDARIA"]).pack()
        tk.Label(cab,
                 text=(f"{EMITENTE['endereco']} - {EMITENTE['bairro']} - "
                       f"{EMITENTE['cidade']}/{EMITENTE['uf']} - CEP {EMITENTE['cep']}  |  "
                       f"CNPJ: {EMITENTE['cnpj']}  |  IE: {EMITENTE['inscricao_estadual']}"),
                 font=("Calibri",7), fg=CORES["ACENTO"], bg=CORES["SECUNDARIA"]).pack()
        tk.Label(cab, text="ASSISTENTE FISCAL - NF-e / NFS-e",
                 font=("Calibri",10,"bold"), fg=CORES["FUNDO_CLARO"], bg=CORES["SECUNDARIA"]).pack(pady=(3,0))

        fp = tk.Frame(self.root, bg=CORES["FUNDO_MEDIO"], pady=5)
        fp.pack(fill="x", padx=20, pady=(8,0))
        tk.Label(fp, text="Progresso:", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_MEDIO"], fg=CORES["SECUNDARIA"]).pack(anchor="w")
        self.barra = ttk.Progressbar(fp, orient="horizontal",
                                     length=760, mode="determinate", maximum=7)
        self.barra.pack(pady=3)
        self.lbl_etapa = tk.Label(fp, text="Etapa 0 de 7",
                                  font=("Calibri",8), bg=CORES["FUNDO_MEDIO"], fg=CORES["TEXTO_ESCURO"])
        self.lbl_etapa.pack(anchor="e")

        self.area = tk.Frame(self.root, bg=CORES["FUNDO_CLARO"], padx=20, pady=10)
        self.area.pack(fill="both", expand=True)

        rod = tk.Frame(self.root, bg=CORES["SECUNDARIA"], pady=5)
        rod.pack(fill="x", side="bottom")
        tk.Label(rod, text="Confirme NCM e CFOP com seu contador antes de emitir.",
                 font=("Calibri",8,"italic"), fg=CORES["ACENTO"], bg=CORES["SECUNDARIA"]).pack()

    def _prog(self, n):
        self.barra["value"] = n
        self.lbl_etapa.config(text=f"Etapa {n} de 7  |  {'▓'*n}{'░'*(7-n)}")

    def _limpar(self):
        for w in self.area.winfo_children():
            w.destroy()

    def _titulo(self, txt, cor):
        f = tk.Frame(self.area, bg=cor, pady=7)
        f.pack(fill="x", pady=(0,8))
        tk.Label(f, text=txt, font=("Calibri",11,"bold"),
                 fg=CORES["TEXTO_CLARO"], bg=cor, padx=10).pack(anchor="w")

    def _btn_nav(self, cmd_avancar, cmd_voltar=None, txt_avancar="  CONTINUAR  ▶"):
        btn_frame = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"])
        btn_frame.pack(pady=12)

        if cmd_voltar:
            tk.Button(btn_frame, text="  ◀ VOLTAR  ", font=("Calibri",11,"bold"),
                      bg=CORES["TEXTO_ESCURO"], fg=CORES["TEXTO_CLARO"], relief="flat",
                      padx=18, pady=8, cursor="hand2",
                      command=cmd_voltar).pack(side="left", padx=5)

        tk.Button(btn_frame, text=txt_avancar, font=("Calibri",11,"bold"),
                  bg=CORES["PRIMARIA"], fg=CORES["TEXTO_CLARO"], relief="flat",
                  padx=18, pady=8, cursor="hand2",
                  command=cmd_avancar).pack(side="left", padx=5)

    def _entry(self, parent, label, ph, width=46):
        f = tk.Frame(parent, bg=CORES["FUNDO_CLARO"]); f.pack(fill="x", pady=4)
        tk.Label(f, text=label, font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"], width=28, anchor="w").pack(side="left")
        e = tk.Entry(f, font=("Calibri",10), width=width, relief="groove", bd=2,
                     highlightbackground=CORES["BORDA"], highlightthickness=1)
        e.insert(0, ph); e.config(fg=CORES["BORDA"])
        e.bind("<FocusIn>",  lambda ev, p=ph, en=e:
               (en.delete(0,"end"), en.config(fg=CORES["TEXTO_ESCURO"])) if en.get()==p else None)
        e.bind("<FocusOut>", lambda ev, p=ph, en=e:
               (en.insert(0,p), en.config(fg=CORES["BORDA"])) if en.get()=="" else None)
        e.pack(side="left", padx=6)
        return e

    def _avancar_etapa(self):
        if self.etapa_atual < len(self.etapas) - 1:
            self.etapa_atual += 1
            self.etapas[self.etapa_atual]()
            self._prog(self.etapa_atual)

    def _voltar_etapa(self):
        if self.etapa_atual > 0:
            self.etapa_atual -= 1
            self.etapas[self.etapa_atual]()
            self._prog(self.etapa_atual)

    def _salvar_e_avancar(self, chave, valor):
        self.dados[chave] = valor
        self._avancar_etapa()

    # --- Etapas do Assistente ---
    def etapa_boas_vindas(self):
        self._limpar()
        tk.Label(self.area, text="Bem-vindo ao Assistente Fiscal!",
                 font=("Calibri",13,"bold"), bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(pady=(15,6))
        tk.Label(self.area,
                 text=("Preencha os 7 blocos para gerar o rascunho da nota:\n\n"
                       "  Bloco 1 - Tipo de Nota Fiscal\n"
                       "  Bloco 2 - Dados do Destinatario\n"
                       "  Bloco 3 - Dados do Equipamento\n"
                       "  Bloco 4 - Codigos Fiscais (NCM / CFOP)\n"
                       "  Bloco 5 - Condicoes da Operacao\n"
                       "  Bloco 6 - Observacoes\n"
                       "  Bloco 7 - Resumo Final\n\n"
                       "Ao final exporte em Excel, PDF (DANFE), HTML ou envie para API."),
                 font=("Calibri",10), bg=CORES["FUNDO_CLARO"], fg=CORES["TEXTO_ESCURO"], justify="left").pack(pady=6)
        tk.Label(self.area,
                 text="ATENCAO: Confirme NCM e CFOP com seu contador!",
                 font=("Calibri",10,"bold"), bg=CORES["FUNDO_ALERTA"], fg=CORES["AVISO"],
                 relief="groove", padx=10, pady=6).pack(pady=8, fill="x")
        self._btn_nav(self._avancar_etapa, None, "  INICIAR PREENCHIMENTO  ▶")

    def etapa_tipo_nota(self):
        self._limpar()
        self._titulo("BLOCO 1 - TIPO DE NOTA FISCAL", CORES["PRIMARIA"])
        tk.Label(self.area, text="Selecione o tipo de operacao:",
                 font=("Calibri",10), bg=CORES["FUNDO_CLARO"], fg=CORES["TEXTO_ESCURO"]).pack(anchor="w", pady=(8,4))
        self.v_tipo = tk.StringVar(value=self.dados.get("tipo_nota", "NF-e - Venda de Equipamento / Produto"))
        for op in ["NF-e - Venda de Equipamento / Produto",
                   "NFS-e - Prestacao de Servico",
                   "NF-e - Devolucao de Mercadoria",
                   "NF-e - Remessa / Demonstracao",
                   "NF-e - Transferencia entre Filiais",
                   "NF-e - Bonificacao / Brinde"]:
            tk.Radiobutton(self.area, text=op, variable=self.v_tipo, value=op,
                           font=("Calibri",10), bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"],
                           activebackground=CORES["FUNDO_CLARO"], selectcolor=CORES["PRIMARIA"]
                           ).pack(anchor="w", pady=2, padx=20)
        self._btn_nav(lambda: self._salvar_e_avancar("tipo_nota", self.v_tipo.get()), self._voltar_etapa)

    def etapa_destinatario(self):
        self._limpar()
        self._titulo("BLOCO 2 - DADOS DO DESTINATARIO", CORES["PRIMARIA"])
        ff = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"]); ff.pack(fill="both", expand=True)
        specs = [
            ("razao_social",  "Razao Social / Nome *",  "Ex: Empresa ABC Ltda"),
            ("cpf_cnpj",      "CPF / CNPJ *",           "Ex: 00.000.000/0001-00"),
            ("insc_estadual", "Inscricao Estadual",      "ISENTO se aplicavel"),
            ("endereco",      "Endereco Completo *",     "Rua, numero, bairro, cidade, UF, CEP"),
            ("email",         "E-mail *",                "financeiro@empresa.com.br"),
        ]
        self.e_dest = {}
        for k, lb, ph in specs:
            val_atual = self.dados.get(k, ph)
            entry = self._entry(ff, lb, val_atual if val_atual != "Nao informado" else ph)
            self.e_dest[k] = entry
        self._btn_nav(self._salvar_dest, self._voltar_etapa)

    def _salvar_dest(self):
        phs = {"razao_social":"Ex: Empresa ABC Ltda","cpf_cnpj":"Ex: 00.000.000/0001-00",
               "insc_estadual":"ISENTO se aplicavel","endereco":"Rua, numero, bairro, cidade, UF, CEP",
               "email":"financeiro@empresa.com.br"}
        for k, e in self.e_dest.items():
            v = e.get().strip()
            self.dados[k] = "Nao informado" if (v == phs.get(k,"") or v == "") else v
        falt = [k for k in ["razao_social","cpf_cnpj","endereco","email"]
                if self.dados[k] == "Nao informado"]
        if falt:
            messagebox.showwarning("Obrigatorios",
                "Preencha:\n" + "\n".join(f"- {k.replace('_',' ').title()}" for k in falt))
            return
        self._avancar_etapa()

    def etapa_equipamento(self):
        self._limpar()
        self._titulo("BLOCO 3 - DADOS DO EQUIPAMENTO", CORES["PRIMARIA"])
        ff = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"]); ff.pack(fill="both", expand=True, pady=6)
        tk.Label(ff, text="Descricao Completa *", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w")
        self.txt_desc = tk.Text(ff, font=("Calibri",10), height=4, width=80,
                                relief="groove", bd=2, highlightbackground=CORES["BORDA"], highlightthickness=1)
        desc_val = self.dados.get("descricao", "Ex: Impressora Laser Multifuncional Marca XYZ Modelo ABC-500")
        self.txt_desc.insert("1.0", desc_val if desc_val != "Nao informado" else "Ex: Impressora Laser Multifuncional Marca XYZ Modelo ABC-500")
        self.txt_desc.config(fg=CORES["TEXTO_ESCURO"] if desc_val != "Nao informado" else CORES["BORDA"])
        self.txt_desc.bind("<FocusIn>", lambda e: (
            self.txt_desc.delete("1.0","end"),
            self.txt_desc.config(fg=CORES["TEXTO_ESCURO"])
        ) if self.txt_desc.get("1.0","end-1c").startswith("Ex:") else None)
        self.txt_desc.pack(pady=6)

        self.e_qtd = self._entry(ff, "Quantidade *", "1", width=10)
        self.e_qtd.pack_forget() # Remove o pack anterior
        self.e_qtd.pack(anchor="w", pady=4, padx=4) # Repack com anchor

        self.e_serie = self._entry(ff, "Numero de Serie", "Nao informado", width=30)
        self.e_serie.pack_forget()
        self.e_serie.pack(anchor="w", pady=4, padx=4)

        self.e_vlr_unit = self._entry(ff, "Valor Unitario (R$) *", "0.00", width=15)
        self.e_vlr_unit.pack_forget()
        self.e_vlr_unit.pack(anchor="w", pady=4, padx=4)

        self._btn_nav(self._salvar_equipamento, self._voltar_etapa)

    def _salvar_equipamento(self):
        desc = self.txt_desc.get("1.0","end-1c").strip()
        self.dados["descricao"] = desc if (desc and not desc.startswith("Ex:")) else "Nao informado"
        self.dados["quantidade"] = self.e_qtd.get().strip()
        self.dados["num_serie"] = self.e_serie.get().strip()
        self.dados["valor_unit"] = self.e_vlr_unit.get().strip()

        try:
            qtd = float(self.dados["quantidade"].replace(",", "."))
            vlr_unit = float(self.dados["valor_unit"].replace("R$", "").replace(".", "").replace(",", ".").strip())
            self.dados["valor_total"] = f"R$ {qtd * vlr_unit:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except ValueError:
            self.dados["valor_total"] = "Verificar"

        falt = [k for k in ["descricao","quantidade","valor_unit"]
                if self.dados[k] in ("Nao informado","","0.00")]
        if falt:
            messagebox.showwarning("Obrigatorios",
                "Preencha:\n" + "\n".join(f"- {k.replace('_',' ').title()}" for k in falt))
            return
        self._avancar_etapa()

    def etapa_codigos_fiscais(self):
        self._limpar()
        self._titulo("BLOCO 4 - CODIGOS FISCAIS (NCM / CFOP)", CORES["ERRO"])
        tk.Label(self.area, text="ATENCAO: Confirme NCM e CFOP com seu contador!",
                 font=("Calibri",9,"bold"), bg=CORES["FUNDO_ALERTA"], fg=CORES["AVISO"],
                 relief="groove", padx=8, pady=5).pack(fill="x", pady=(4,10))
        ff = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"]); ff.pack(fill="both", expand=True)

        # NCM
        tk.Label(ff, text="Codigo NCM (8 digitos) *", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w")
        tk.Label(ff, text="8471.30.19 Computadores | 8443.31.99 Impressoras | 8517.62.99 Roteadores | 8528.72.00 Monitores",
                 font=("Calibri",7), bg=CORES["FUNDO_CLARO"], fg=CORES["TEXTO_ESCURO"]).pack(anchor="w", padx=4)
        self.e_ncm = self._entry(ff, "", self.dados.get("ncm", "Ex: 84713019"), width=20)
        self.e_ncm.pack_forget()
        self.e_ncm.pack(anchor="w", pady=(2,10), padx=4)

        # CFOP
        tk.Label(ff, text="CFOP (Codigo Fiscal de Operacoes e Prestacoes) *", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w")

        # Inicializa self.v_cfop ANTES de _atualizar_cfop_opcoes
        cfop_salvo = self.dados.get("cfop", "")
        if cfop_salvo and " - " not in cfop_salvo: # Se for um CFOP digitado manualmente
            self.v_cfop = tk.StringVar(value="outro - Outro (informar manualmente)")
            self.cfop_manual_inicial = cfop_salvo # Guarda o valor para preencher o campo manual
        else:
            self.v_cfop = tk.StringVar(value=cfop_salvo if cfop_salvo else "")
            self.cfop_manual_inicial = ""

        self.cfop_radio_frame = tk.Frame(ff, bg=CORES["FUNDO_CLARO"])
        self.cfop_radio_frame.pack(anchor="w", pady=2, padx=4)

        self.e_cfop_manual = self._entry(ff, "CFOP Manual", self.cfop_manual_inicial if self.cfop_manual_inicial else "Digite o CFOP aqui", width=30)
        self.e_cfop_manual.pack_forget() # Esconde inicialmente

        # Vincula o NCM e atualiza as opcoes de CFOP
        self.e_ncm.bind("<KeyRelease>", self._sugerir_cfop)
        self._atualizar_cfop_opcoes() # Chama para popular as opções iniciais de CFOP

        # Natureza da Operacao
        tk.Label(ff, text="Natureza da Operacao *", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w", pady=(8,0))
        self.cb_nat = ttk.Combobox(ff, font=("Calibri",10), width=45,
                                   values=["VENDA DE EQUIPAMENTO","VENDA DE MERCADORIA",
                                           "PRESTACAO DE SERVICO","REMESSA PARA DEMONSTRACAO",
                                           "DEVOLUCAO DE VENDA","TRANSFERENCIA ENTRE FILIAIS",
                                           "EXPORTACAO DE PRODUTO"])
        self.cb_nat.set(self.dados.get("natureza", "VENDA DE EQUIPAMENTO"))
        self.cb_nat.pack(anchor="w", pady=2, padx=4)
        self._btn_nav(self._salvar_codigos, self._voltar_etapa)

    def _sugerir_cfop(self, event=None):
        ncm_digitado = self.e_ncm.get().strip().replace(".","")
        self._atualizar_cfop_opcoes(ncm_digitado)

    def _atualizar_cfop_opcoes(self, ncm_digitado=""):
        for widget in self.cfop_radio_frame.winfo_children():
            widget.destroy()

        opcoes_cfop = NCM_CFOP_MAP.get(ncm_digitado, [])
        if not opcoes_cfop:
            opcoes_cfop = ["Nenhum CFOP sugerido para este NCM"]

        # Adiciona a opção "Outro" se não estiver presente
        if "outro - Outro (informar manualmente)" not in opcoes_cfop:
            opcoes_cfop.append("outro - Outro (informar manualmente)")

        for op in opcoes_cfop:
            rb = tk.Radiobutton(
                self.cfop_radio_frame,
                text=op,
                variable=self.v_cfop,
                value=op,
                font=("Calibri",10),
                bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"],
                activebackground=CORES["FUNDO_CLARO"], selectcolor=CORES["PRIMARIA"],
                command=self._toggle_cfop_manual
            )
            rb.pack(anchor="w", pady=2)

        # Garante que um valor inicial seja selecionado
        if self.v_cfop.get() not in opcoes_cfop and self.v_cfop.get() != "outro - Outro (informar manualmente)":
            self.v_cfop.set(opcoes_cfop[0])

        self._toggle_cfop_manual()

    def _toggle_cfop_manual(self):
        if hasattr(self, 'e_cfop_manual'):
            if self.v_cfop.get() == "outro - Outro (informar manualmente)":
                self.e_cfop_manual.pack(anchor="w", pady=4, padx=4)
                self.e_cfop_manual.config(state="normal", fg=CORES["TEXTO_ESCURO"])
                if self.e_cfop_manual.get() == "Digite o CFOP aqui":
                    self.e_cfop_manual.delete(0, "end")
            else:
                self.e_cfop_manual.pack_forget()
                self.e_cfop_manual.config(state="disabled", fg=CORES["BORDA"])
                self.e_cfop_manual.delete(0, "end")
                self.e_cfop_manual.insert(0, "Digite o CFOP aqui")

    def _salvar_codigos(self):
        ncm = self.e_ncm.get().strip()
        if not ncm or ncm == "Ex: 84713019":
            messagebox.showwarning("Obrigatorio", "Informe o NCM.")
            return
        self.dados["ncm"] = ncm

        cfop_selecionado = self.v_cfop.get()
        if cfop_selecionado == "outro - Outro (informar manualmente)":
            cfop_manual = self.e_cfop_manual.get().strip()
            if not cfop_manual or cfop_manual == "Digite o CFOP aqui":
                messagebox.showwarning("Obrigatorio", "Informe o CFOP manual.")
                return
            self.dados["cfop"] = cfop_manual
        elif cfop_selecionado == "Nenhum CFOP sugerido para este NCM":
            messagebox.showwarning("Obrigatorio", "Selecione um CFOP ou digite manualmente.")
            return
        else:
            self.dados["cfop"] = cfop_selecionado

        natureza = self.cb_nat.get().strip()
        if not natureza:
            messagebox.showwarning("Obrigatorio", "Informe a Natureza da Operacao.")
            return
        self.dados["natureza"] = natureza
        self._avancar_etapa()

    def etapa_condicoes(self):
        self._limpar()
        self._titulo("BLOCO 5 - CONDICOES DA OPERACAO", CORES["PRIMARIA"])
        ff = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"]); ff.pack(fill="both", expand=True)

        # Pagamento
        tk.Label(ff, text="Forma de Pagamento *", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w")
        self.v_pagamento = tk.StringVar(value=self.dados.get("pagamento", "A Vista"))
        for op in ["A Vista", "A Prazo", "Outros"]:
            tk.Radiobutton(ff, text=op, variable=self.v_pagamento, value=op,
                           font=("Calibri",10), bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"],
                           activebackground=CORES["FUNDO_CLARO"], selectcolor=CORES["PRIMARIA"]
                           ).pack(anchor="w", pady=2, padx=20)

        # Frete
        tk.Label(ff, text="Modalidade do Frete *", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w", pady=(10,0))
        self.v_frete = tk.StringVar(value=self.dados.get("frete", "0 - Emitente"))
        for op in ["0 - Emitente", "1 - Destinatario", "2 - Terceiros", "9 - Sem Frete"]:
            tk.Radiobutton(ff, text=op, variable=self.v_frete, value=op,
                           font=("Calibri",10), bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"],
                           activebackground=CORES["FUNDO_CLARO"], selectcolor=CORES["PRIMARIA"]
                           ).pack(anchor="w", pady=2, padx=20)

        # Transportadora
        tk.Label(ff, text="Transportadora", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w", pady=(10,0))
        self.cb_transp = ttk.Combobox(ff, font=("Calibri",10), width=45,
                                      values=["Correios", "Jadlog", "Loggi", "Outra"])
        self.cb_transp.set(self.dados.get("transportadora", "Outra"))
        self.cb_transp.pack(anchor="w", pady=2, padx=4)
        self.e_transp_manual = self._entry(ff, "Nome da Transportadora", self.dados.get("transportadora", "Digite o nome da transportadora"), width=30)
        self.e_transp_manual.pack_forget() # Esconde inicialmente

        # Bind para mostrar/esconder campo manual
        self.cb_transp.bind("<<ComboboxSelected>>", self._toggle_transportadora_manual)
        # Chama uma vez para configurar o estado inicial
        self._toggle_transportadora_manual()

        self._btn_nav(self._salvar_condicoes, self._voltar_etapa)

    def _toggle_transportadora_manual(self, event=None):
        if self.cb_transp.get() == "Outra":
            self.e_transp_manual.pack(anchor="w", pady=4, padx=4)
            self.e_transp_manual.config(state="normal", fg=CORES["TEXTO_ESCURO"])
            if self.e_transp_manual.get() == "Digite o nome da transportadora":
                self.e_transp_manual.delete(0, "end")
        else:
            self.e_transp_manual.pack_forget()
            self.e_transp_manual.config(state="disabled", fg=CORES["BORDA"])
            self.e_transp_manual.delete(0, "end")
            self.e_transp_manual.insert(0, "Digite o nome da transportadora")

    def _salvar_condicoes(self):
        self.dados["pagamento"] = self.v_pagamento.get()
        self.dados["frete"] = self.v_frete.get()
        if self.cb_transp.get() == "Outra":
            transp_manual = self.e_transp_manual.get().strip()
            self.dados["transportadora"] = transp_manual if transp_manual and transp_manual != "Digite o nome da transportadora" else "Nao informado"
        else:
            self.dados["transportadora"] = self.cb_transp.get()
        self._avancar_etapa()

    def etapa_observacoes(self):
        self._limpar()
        self._titulo("BLOCO 6 - OBSERVACOES", CORES["PRIMARIA"])
        ff = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"]); ff.pack(fill="both", expand=True, pady=6)
        tk.Label(ff, text="Observacoes Adicionais", font=("Calibri",9,"bold"),
                 bg=CORES["FUNDO_CLARO"], fg=CORES["SECUNDARIA"]).pack(anchor="w")
        self.txt_obs = tk.Text(ff, font=("Calibri",10), height=8, width=80,
                               relief="groove", bd=2, highlightbackground=CORES["BORDA"], highlightthickness=1)
        obs_val = self.dados.get("observacoes", "Sem observacoes")
        self.txt_obs.insert("1.0", obs_val if obs_val != "Sem observacoes" else "Sem observacoes")
        self.txt_obs.config(fg=CORES["TEXTO_ESCURO"] if obs_val != "Sem observacoes" else CORES["BORDA"])
        self.txt_obs.bind("<FocusIn>", lambda e: (
            self.txt_obs.delete("1.0","end"),
            self.txt_obs.config(fg=CORES["TEXTO_ESCURO"])
        ) if self.txt_obs.get("1.0","end-1c").startswith("Sem observacoes") else None)
        self.txt_obs.pack(pady=6)
        self._btn_nav(self._salvar_observacoes, self._voltar_etapa)

    def _salvar_observacoes(self):
        obs = self.txt_obs.get("1.0","end-1c").strip()
        self.dados["observacoes"] = obs if (obs and not obs.startswith("Sem observacoes")) else "Sem observacoes"
        self._avancar_etapa()

    def etapa_resumo_final(self):
        self._limpar()
        self._titulo("BLOCO 7 - RESUMO FINAL", CORES["SUCESSO"])

        resumo_frame = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"])
        resumo_frame.pack(fill="both", expand=True, pady=10)

        # Exibir dados em um Text widget ou Labels
        resumo_text = tk.Text(resumo_frame, wrap="word", font=("Calibri", 10),
                              bg=CORES["FUNDO_CLARO"], fg=CORES["TEXTO_ESCURO"],
                              relief="flat", height=20)
        resumo_text.pack(fill="both", expand=True)
        resumo_text.insert("end", "--- DADOS DA NOTA FISCAL ---\n\n")
        for k, v in self.dados.items():
            resumo_text.insert("end", f"{k.replace('_', ' ').title()}: {v}\n")
        resumo_text.config(state="disabled") # Torna o texto somente leitura

        btn_frame = tk.Frame(self.area, bg=CORES["FUNDO_CLARO"])
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="  EXPORTAR EXCEL  ", font=("Calibri",10,"bold"),
                  bg=CORES["SUCESSO"], fg=CORES["TEXTO_CLARO"], relief="flat",
                  padx=15, pady=7, cursor="hand2",
                  command=self.exportar_excel).pack(side="left", padx=5)
        tk.Button(btn_frame, text="  EXPORTAR DANFE PDF  ", font=("Calibri",10,"bold"),
                  bg=CORES["SUCESSO"], fg=CORES["TEXTO_CLARO"], relief="flat",
                  padx=15, pady=7, cursor="hand2",
                  command=self.exportar_pdf).pack(side="left", padx=5)
        tk.Button(btn_frame, text="  VISUALIZAR HTML  ", font=("Calibri",10,"bold"),
                  bg=CORES["SUCESSO"], fg=CORES["TEXTO_CLARO"], relief="flat",
                  padx=15, pady=7, cursor="hand2",
                  command=self.exportar_html).pack(side="left", padx=5)
        tk.Button(btn_frame, text="  ENVIAR PARA API  ", font=("Calibri",10,"bold"),
                  bg=CORES["SUCESSO"], fg=CORES["TEXTO_CLARO"], relief="flat",
                  padx=15, pady=7, cursor="hand2",
                  command=self.enviar_para_api).pack(side="left", padx=5)

        self._btn_nav(self.nova_nota, self._voltar_etapa, "  NOVA NOTA  ")

    def exportar_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Ausente","A biblioteca 'openpyxl' nao esta instalada. Por favor, instale com 'pip install openpyxl'."); return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rascunho NF"

            # Estilos
            fn = lambda b, s, c: Font(name="Calibri", size=s, bold=b, color=c)
            fl = lambda c: PatternFill(start_color=c, end_color=c, fill_type="solid")
            ale_wrap = lambda: Alignment(horizontal="left", vertical="center", wrap_text=True)
            bd = lambda: Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Cabecalho
            ws.merge_cells("B2:G2")
            ws["B2"] = EMITENTE["razao_social"]
            ws["B2"].font = fn(True, 14, CORES["SECUNDARIA"].replace("#",""))
            ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 25

            ws.merge_cells("B3:G3")
            ws["B3"] = (f"{EMITENTE['endereco']} - {EMITENTE['bairro']} - "
                        f"{EMITENTE['cidade']}/{EMITENTE['uf']} - CEP {EMITENTE['cep']}  |  "
                        f"CNPJ: {EMITENTE['cnpj']}  |  IE: {EMITENTE['inscricao_estadual']}")
            ws["B3"].font = fn(False, 8, CORES["TEXTO_ESCURO"].replace("#",""))
            ws["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[3].height = 20

            ws.merge_cells("B5:G5")
            ws["B5"] = "RASCUNHO DE NOTA FISCAL - SEM VALOR FISCAL"
            ws["B5"].font = fn(True, 12, CORES["ERRO"].replace("#",""))
            ws["B5"].fill = fl(CORES["FUNDO_ALERTA"].replace("#",""))
            ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[5].height = 20

            linha = 7
            alt = 0 # Contador para linhas alternadas

            # Largura das colunas
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 50

            secoes = [
                ("NATUREZA DA OPERACAO", CORES["PRIMARIA"], [("Tipo de Nota", self.dados["tipo_nota"]),
                                              ("Natureza", self.dados["natureza"]),
                                              ("CFOP", self.dados["cfop"]),
                                              ("NCM", self.dados["ncm"])]),
                ("DESTINATARIO", CORES["PRIMARIA"], [("Razao Social / Nome", self.dados["razao_social"]),
                                      ("CPF / CNPJ", self.dados["cpf_cnpj"]),
                                      ("Inscricao Estadual", self.dados["insc_estadual"]),
                                      ("Endereco", self.dados["endereco"]),
                                      ("E-mail", self.dados["email"])]),
                ("DADOS DO PRODUTO / EQUIPAMENTO", CORES["PRIMARIA"], [("Descricao", self.dados["descricao"]),
                                                        ("Quantidade", self.dados["quantidade"]),
                                                        ("Valor Unitario", self.dados["valor_unit"]),
                                                        ("Valor Total", self.dados["valor_total"]),
                                                        ("Numero de Serie", self.dados["num_serie"])]),
                ("CALCULO DO IMPOSTO E TOTAIS", CORES["PRIMARIA"], [("BC ICMS", "A calcular no sistema"),
                                                      ("Valor ICMS", "A calcular no sistema"),
                                                      ("BC ICMS ST", "A calcular no sistema"),
                                                      ("Valor ICMS ST", "A calcular no sistema"),
                                                      ("Valor Produtos", self.dados["valor_total"]),
                                                      ("Valor Frete", "Verificar"),
                                                      ("Valor Seguro", "0,00"),
                                                      ("Desconto", "0,00"),
                                                      ("Outras Despesas", "0,00"),
                                                      ("Valor IPI", "A calcular no sistema"),
                                                      ("Valor PIS", "A calcular no sistema"),
                                                      ("Valor COFINS", "A calcular no sistema"),
                                                      ("VALOR TOTAL NF", self.dados["valor_total"]),
                                                      ("Forma Pgto", self.dados["pagamento"])]),
                ("TRANSPORTE", CORES["PRIMARIA"],[("Modalidade do Frete", self.dados["frete"]),
                                              ("Transportadora", self.dados["transportadora"]),
                                              ("Qtd / Especie", "1 / EQUIPAMENTO")]),
                ("OBSERVACOES", CORES["PRIMARIA"],[("Observacoes", self.dados["observacoes"])]),
            ]
            for ts, cs, itens in secoes:
                ws.merge_cells(f"B{linha}:G{linha}") # Mescla de B a G para o titulo da secao
                ws[f"B{linha}"] = ts
                ws[f"B{linha}"].font=fn(True,9,CORES["TEXTO_CLARO"].replace("#","")); ws[f"B{linha}"].fill=fl(cs)
                ws[f"B{linha}"].alignment=ale_wrap(); ws.row_dimensions[linha].height=20
                linha += 1
                for campo, valor in itens:
                    bg = CORES["FUNDO_MEDIO"].replace("#","") if alt%2==0 else CORES["FUNDO_CLARO"].replace("#",""); alt+=1
                    ws[f"B{linha}"]=campo
                    ws[f"B{linha}"].font=fn(True,9,CORES["SECUNDARIA"].replace("#",""))
                    ws[f"B{linha}"].fill=fl(bg); ws[f"B{linha}"].alignment=ale_wrap()
                    ws[f"B{linha}"].border=bd()
                    ws.merge_cells(f"C{linha}:G{linha}") # Mescla de C a G para o valor
                    ws[f"C{linha}"]=valor
                    ws[f"C{linha}"].font=fn(campo=="VALOR TOTAL NF",9,
                                           CORES["SUCESSO"].replace("#","") if campo=="VALOR TOTAL NF" else CORES["TEXTO_ESCURO"].replace("#",""))
                    ws[f"C{linha}"].fill=fl(bg); ws[f"C{linha}"].alignment=ale_wrap()
                    ws[f"C{linha}"].border=bd()
                    ws.row_dimensions[linha].height=18; linha+=1
                linha+=1 # Espaco entre secoes

            nome = "".join(c for c in
                f"NF_{self.dados['razao_social'][:12].replace(' ','_')}"
                f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                if c.isalnum() or c in("_","-","."))
            cam = os.path.join(os.path.expanduser("~"),"Desktop",nome)
            try: wb.save(cam)
            except Exception: cam=os.path.join(os.getcwd(),nome); wb.save(cam)
            messagebox.showinfo("Excel Exportado", f"Salvo em:\n\n{cam}")
        except Exception as ex:
            messagebox.showerror("Erro Excel", f"Ocorreu um erro ao exportar para Excel: {str(ex)}")

    def exportar_pdf(self):
        if not PDF_OK:
            messagebox.showerror("Ausente","A biblioteca 'fpdf2' nao esta instalada. Por favor, instale com 'pip install fpdf2'."); return
        try:
            nome = "".join(c for c in
                f"DANFE_{self.dados['razao_social'][:12].replace(' ','_')}"
                f"_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                if c.isalnum() or c in("_","-","."))
            cam = os.path.join(os.path.expanduser("~"),"Desktop",nome)
            GeradorDANFE(self.dados).gerar(cam)
            if messagebox.askyesno("PDF Gerado", f"Salvo em:\n\n{cam}\n\nAbrir agora?"):
                webbrowser.open(f"file://{os.path.abspath(cam)}")
        except Exception as ex:
            messagebox.showerror("Erro PDF", f"Ocorreu um erro ao gerar o PDF: {str(ex)}")

    def exportar_html(self):
        try:
            nome = "".join(c for c in
                f"DANFE_{self.dados['razao_social'][:12].replace(' ','_')}"
                f"_{datetime.now().strftime('%Y%m%d_%H%M')}.html"
                if c.isalnum() or c in("_","-","."))
            cam = os.path.join(os.path.expanduser("~"),"Desktop",nome)
            html_content = GeradorDANFE(self.dados)._montar_html()
            with open(cam, "w", encoding="utf-8") as f:
                f.write(html_content)
            if messagebox.askyesno("HTML Gerado", f"Salvo em:\n\n{cam}\n\nAbrir agora?"):
                webbrowser.open(f"file://{os.path.abspath(cam)}")
        except Exception as ex:
            messagebox.showerror("Erro HTML", f"Ocorreu um erro ao gerar o HTML: {str(ex)}")

    def enviar_para_api(self):
        if not REQUESTS_OK:
            messagebox.showerror("Ausente","A biblioteca 'requests' nao esta instalada. Por favor, instale com 'pip install requests'."); return

        api_url = "https://sua-api.com/endpoint-de-notas"
        headers = {"Content-Type": "application/json"}

        dados_para_api = {k: v for k, v in self.dados.items() if v != "Nao informado" and v != "Sem observacoes"}

        try:
            dados_para_api['quantidade'] = float(dados_para_api.get('quantidade', '0').replace(',', '.'))
            dados_para_api['valor_unit'] = float(dados_para_api.get('valor_unit', '0').replace('R$', '').replace('.', '').replace(',', '.').strip())
            dados_para_api['valor_total'] = float(dados_para_api.get('valor_total', '0').replace('R$', '').replace('.', '').replace(',', '.').strip())
        except ValueError:
            pass

        try:
            response = requests.post(api_url, headers=headers, data=json.dumps(dados_para_api))
            response.raise_for_status()

            messagebox.showinfo("Envio API", f"Dados enviados com sucesso para a API!\nStatus: {response.status_code}\nResposta: {response.json()}")
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Erro API", f"Erro ao enviar dados para a API: {e}\nVerifique a URL e sua conexão.")
        except json.JSONDecodeError:
            messagebox.showerror("Erro API", f"Erro ao decodificar a resposta da API. Resposta bruta: {response.text}")
        except Exception as e:
            messagebox.showerror("Erro API", f"Ocorreu um erro inesperado: {e}")

    def nova_nota(self):
        if messagebox.askyesno("Nova Nota","Iniciar nova nota? Dados atuais serao apagados."):
            self.dados = dict(self.DADOS_VAZIOS)
            self.etapa_atual = 0; self._prog(0); self.etapa_boas_vindas()


# ================================================================
# PONTO DE ENTRADA
# ================================================================
def main():
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "NF_ASSIST.html")
    if os.path.exists(html_path):
        print(f"Abrindo assistente no navegador: {html_path}")
        webbrowser.open(f"file:///{html_path.replace(os.sep, '/')}")
        return

    # Fallback: interface Tkinter (caso o HTML nao seja encontrado)
    if not EXCEL_OK: print("AVISO: pip install openpyxl")
    if not PDF_OK:   print("AVISO: pip install fpdf2")
    if not REQUESTS_OK: print("AVISO: pip install requests")
    root = tk.Tk()
    estilo = ttk.Style()
    estilo.theme_use("clam")
    estilo.configure("TProgressbar",
                     troughcolor=CORES["FUNDO_MEDIO"],
                     background=CORES["PRIMARIA"],
                     thickness=12)
    AssistenteFiscal(root)
    root.mainloop()

if __name__ == "__main__":
    main()